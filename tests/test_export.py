"""Excel export tests: table helper (unit) and full workbook build (integration)."""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook, load_workbook

from claimsight.export.excel_report import _write_table


def test_write_table_creates_excel_table():
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
    ref = _write_table(ws, df, "T1")
    assert ref == "A1:B3"
    assert "T1" in ws.tables
    assert ws["A1"].value == "a"
    assert ws["A2"].value == 1
    # Frozen header row.
    assert ws.freeze_panes == "A2"


def test_build_workbook_has_expected_sheets(db_engine, tmp_path):
    # Requires the reporting views to exist; skip cleanly if not built yet.
    import sqlalchemy

    from claimsight.config import get_settings
    from claimsight.export.excel_report import build_workbook

    settings = get_settings()
    try:
        with db_engine.connect() as conn:
            conn.execute(
                sqlalchemy.text(f'SELECT 1 FROM "{settings.reporting_schema}".v_claims_enriched LIMIT 1')
            )
    except sqlalchemy.exc.SQLAlchemyError:
        import pytest
        pytest.skip("reporting views not built")

    out = build_workbook(settings=settings, out_path=tmp_path / "report.xlsx")
    wb = load_workbook(out)
    assert {"KPI Summary", "Claims Detail", "Provider Scorecard", "Data Quality"}.issubset(
        set(wb.sheetnames)
    )
