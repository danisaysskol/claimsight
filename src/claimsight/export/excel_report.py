"""Build the Excel executive report with **live formulas** (not static values).

Sheets:
  * KPI Summary       — headline KPIs, named ranges, conditional formatting
  * Claims Detail     — an Excel Table (autofilter + frozen panes) of claims
  * Provider Scorecard— SUMIFS / AVERAGEIFS / COUNTIFS over the Claims Detail
                        table, so the formula work is visible and recalculates
  * Data Quality      — the latest DQ rule results

Number formats: PKR with thousands separators, percentages to one decimal.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

from claimsight.config import EXPORTS_DIR, Settings, get_settings
from claimsight.db import get_engine

PKR_FMT = '#,##0'
PCT_FMT = '0.0%'
HEADER_FILL = PatternFill("solid", fgColor="0072B2")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="0072B2")


def _load(settings: Settings) -> dict[str, pd.DataFrame]:
    engine = get_engine(settings)
    schema = settings.reporting_schema
    with engine.connect() as conn:
        claims = pd.read_sql(f'SELECT * FROM "{schema}".v_claims_enriched', conn)
        fin = pd.read_sql(f'SELECT * FROM "{schema}".v_financial_monthly', conn)
        scorecard = pd.read_sql(f'SELECT * FROM "{schema}".v_provider_scorecard', conn)
        dq = pd.read_sql(
            "SELECT rule_id, rule_name, dimension, severity, rows_checked, rows_failed, "
            "fail_rate, passed FROM dq.dq_results WHERE run_id = "
            "(SELECT max(run_id) FROM dq.dq_results) ORDER BY severity, rule_id",
            conn,
        )
    return {"claims": claims, "fin": fin, "scorecard": scorecard, "dq": dq}


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_table(ws, df: pd.DataFrame, name: str, start_row: int = 1) -> str:
    """Write a DataFrame as an Excel Table; return the table range ref."""
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col))
    for i, (_, r) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = r[col]
            ws.cell(row=i, column=j, value=None if pd.isna(val) else (val.item() if hasattr(val, "item") else val))
    end_row = start_row + len(df)
    end_col = get_column_letter(len(df.columns))
    ref = f"A{start_row}:{end_col}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return ref


def build_workbook(settings: Settings | None = None, out_path: Path | None = None) -> Path:
    settings = settings or get_settings()
    out_path = out_path or (EXPORTS_DIR / "ClaimSight_Executive_Report.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    data = _load(settings)

    wb = Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: KPI Summary
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = "KPI Summary"
    ws["A1"] = "ClaimSight — Executive KPI Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "All amounts in PKR. Synthetic data — no real patient information."
    ws["A2"].font = Font(italic=True, size=9)

    claims = data["claims"]
    total_billed = float(claims["billed_amount_pkr"].sum()) if not claims.empty else 0.0
    total_approved = float(claims["approved_amount_pkr"].sum()) if not claims.empty else 0.0
    total_paid = float(claims["paid_amount_pkr"].sum()) if not claims.empty else 0.0
    claim_count = int(len(claims))

    kpis = [
        ("Total claims", claim_count, PKR_FMT),
        ("Total billed", total_billed, PKR_FMT),
        ("Total approved", total_approved, PKR_FMT),
        ("Total paid", total_paid, PKR_FMT),
    ]
    row0 = 4
    ws.cell(row=row0, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=row0, column=2, value="Value").font = Font(bold=True)
    for i, (label, val, fmt) in enumerate(kpis, start=row0 + 1):
        ws.cell(row=i, column=1, value=label)
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = fmt
    # Live ratio formulas referencing the cells above.
    ws.cell(row=row0 + 5, column=1, value="Approval rate")
    ws.cell(row=row0 + 5, column=2, value=f"=B{row0 + 3}/B{row0 + 2}").number_format = PCT_FMT
    ws.cell(row=row0 + 6, column=1, value="Savings rate")
    ws.cell(row=row0 + 6, column=2, value=f"=(B{row0 + 2}-B{row0 + 3})/B{row0 + 2}").number_format = PCT_FMT

    # Named ranges for the headline figures.
    wb.defined_names["KPI_TotalPaid"] = DefinedName("KPI_TotalPaid", attr_text=f"'KPI Summary'!$B${row0 + 4}")
    wb.defined_names["KPI_TotalBilled"] = DefinedName("KPI_TotalBilled", attr_text=f"'KPI Summary'!$B${row0 + 2}")

    # Monthly trend block with a colour-scale + data-bar conditional format.
    fin = data["fin"]
    tstart = row0 + 9
    ws.cell(row=tstart - 1, column=1, value="Monthly trend").font = Font(bold=True, size=12)
    headers = ["Month", "Billed", "Approved", "Paid", "Approval rate"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=tstart, column=j, value=h)
    _style_header(ws, tstart, len(headers))
    for i, (_, r) in enumerate(fin.iterrows(), start=tstart + 1):
        ws.cell(row=i, column=1, value=str(r["year_month"]))
        ws.cell(row=i, column=2, value=float(r["billed_pkr"] or 0)).number_format = PKR_FMT
        ws.cell(row=i, column=3, value=float(r["approved_pkr"] or 0)).number_format = PKR_FMT
        ws.cell(row=i, column=4, value=float(r["paid_pkr"] or 0)).number_format = PKR_FMT
        ws.cell(row=i, column=5, value=float(r["approval_rate"] or 0)).number_format = PCT_FMT
    tend = tstart + len(fin)
    if len(fin) > 0:
        ws.conditional_formatting.add(
            f"D{tstart + 1}:D{tend}",
            DataBarRule(start_type="min", end_type="max", color="009E73"),
        )
        ws.conditional_formatting.add(
            f"E{tstart + 1}:E{tend}",
            ColorScaleRule(start_type="min", start_color="D55E00",
                           mid_type="percentile", mid_value=50, mid_color="F0E442",
                           end_type="max", end_color="009E73"),
        )
    ws.column_dimensions["A"].width = 24
    for col in "BCDE":
        ws.column_dimensions[col].width = 16

    # ------------------------------------------------------------------ #
    # Sheet 2: Claims Detail (Excel Table)
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet("Claims Detail")
    detail_cols = [
        "claim_id", "claim_date", "employer_name", "provider_id", "hospital_name",
        "member_city", "network_status", "claim_type", "status", "adjudication_mode",
        "billed_amount_pkr", "approved_amount_pkr", "paid_amount_pkr", "savings_pkr",
    ]
    detail = claims[detail_cols].copy() if not claims.empty else pd.DataFrame(columns=detail_cols)
    detail["claim_date"] = detail["claim_date"].astype(str)
    _write_table(ws2, detail, "ClaimsDetail")
    for j, col in enumerate(detail_cols, start=1):
        if col.endswith("_pkr"):
            for i in range(2, len(detail) + 2):
                ws2.cell(row=i, column=j).number_format = PKR_FMT
    # Column references within the ClaimsDetail table for the scorecard formulas.
    col_letters = {c: get_column_letter(j) for j, c in enumerate(detail_cols, start=1)}
    n = len(detail)

    # ------------------------------------------------------------------ #
    # Sheet 3: Provider Scorecard (LIVE formulas over ClaimsDetail)
    # ------------------------------------------------------------------ #
    ws3 = wb.create_sheet("Provider Scorecard")
    ws3["A1"] = "Provider Scorecard — live formulas over the Claims Detail table"
    ws3["A1"].font = TITLE_FONT
    sc_headers = ["Provider ID", "Hospital", "Claim count", "Total paid (PKR)",
                  "Avg claim value (PKR)", "Total billed (PKR)"]
    hr = 3
    for j, h in enumerate(sc_headers, start=1):
        ws3.cell(row=hr, column=j, value=h)
    _style_header(ws3, hr, len(sc_headers))

    scorecard = data["scorecard"]
    prov_col = col_letters["provider_id"]
    paid_col = col_letters["paid_amount_pkr"]
    billed_col = col_letters["billed_amount_pkr"]
    rng = lambda c: f"'Claims Detail'!${c}$2:${c}${n + 1}"  # noqa: E731
    for i, (_, r) in enumerate(scorecard.head(60).iterrows(), start=hr + 1):
        pid = r["provider_id"]
        ws3.cell(row=i, column=1, value=pid)
        ws3.cell(row=i, column=2, value=r["hospital_name"])
        crit = f'{rng(prov_col)},$A{i}'
        ws3.cell(row=i, column=3, value=f"=COUNTIFS({crit})")
        ws3.cell(row=i, column=4, value=f"=SUMIFS({rng(paid_col)},{crit})").number_format = PKR_FMT
        ws3.cell(row=i, column=5, value=f"=IFERROR(AVERAGEIFS({rng(billed_col)},{crit}),0)").number_format = PKR_FMT
        ws3.cell(row=i, column=6, value=f"=SUMIFS({rng(billed_col)},{crit})").number_format = PKR_FMT
    for col, w in zip("ABCDEF", [14, 30, 14, 18, 20, 18], strict=True):
        ws3.column_dimensions[col].width = w

    # ------------------------------------------------------------------ #
    # Sheet 4: Data Quality
    # ------------------------------------------------------------------ #
    ws4 = wb.create_sheet("Data Quality")
    ws4["A1"] = "Data Quality — latest run"
    ws4["A1"].font = TITLE_FONT
    dq = data["dq"]
    if not dq.empty:
        dq = dq.copy()
        dq["fail_rate"] = dq["fail_rate"].astype(float)
        _write_table(ws4, dq, "DataQuality", start_row=3)
        fr_col = get_column_letter(list(dq.columns).index("fail_rate") + 1)
        for i in range(4, len(dq) + 4):
            ws4.cell(row=i, column=list(dq.columns).index("fail_rate") + 1).number_format = PCT_FMT
        ws4.conditional_formatting.add(
            f"{fr_col}4:{fr_col}{len(dq) + 3}",
            ColorScaleRule(start_type="num", start_value=0, start_color="009E73",
                           end_type="max", end_color="D55E00"),
        )

    wb.save(out_path)
    return out_path


def main() -> None:
    path = build_workbook()
    print(f"Wrote Excel report to {path}")


if __name__ == "__main__":
    main()
